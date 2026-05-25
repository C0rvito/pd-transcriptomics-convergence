import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

class AbntTableFormatter:
    """Formata tabelas para o padrão ABNT (Excel)."""
    
    TRADUCOES = {
        'gene_symbol': 'Símbolo do Gene',
        'symbol': 'Símbolo',
        'log2FoldChange': 'Log2 Fold Change',
        'padj': 'Valor-p Ajustado',
        'pvalue': 'Valor-p',
        'baseMean': 'Média Base',
        'abs_log2FC': 'Abs Log2 FC',
        'UP/DOWN': 'Expressão'
    }

    def __init__(self, font_family="Arial", font_size=10):
        self.font_family = font_family
        self.font_size = font_size

    def formatar_tabela(self, input_path: Path, output_path: Path, title: str):
        """Lê uma tabela (TSV/CSV), traduz e salva em Excel no padrão ABNT."""
        if input_path.suffix == '.tsv':
            df = pd.read_csv(input_path, sep='\t')
        else:
            df = pd.read_csv(input_path)

        # 1. Traduzir Colunas
        df = df.rename(columns=self.TRADUCOES)
        
        # 2. Arredondar valores numéricos
        cols_numericas = df.select_dtypes(include=['float64', 'float32']).columns
        df[cols_numericas] = df[cols_numericas].round(4)

        # 3. Criar Excel e aplicar estilo ABNT
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Tabela')
            
            workbook = writer.book
            worksheet = writer.sheets['Tabela']
            
            # Estilo das Bordas (ABNT: Bordas horizontais apenas no topo e base do cabeçalho e fim da tabela)
            thin = Side(border_style="thin", color="000000")
            header_border = Border(top=thin, bottom=thin)
            bottom_border = Border(bottom=thin)
            
            header_font = Font(name=self.font_family, size=self.font_size, bold=True)
            body_font = Font(name=self.font_family, size=self.font_size)
            
            # Formatar Cabeçalho
            for cell in worksheet[1]:
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal='center')
            
            # Formatar Corpo
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.font = body_font
                    cell.alignment = Alignment(horizontal='center')
                    
            # Borda final na última linha
            for cell in worksheet[worksheet.max_row]:
                cell.border = bottom_border

            # Ajustar largura das colunas
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

        return output_path
